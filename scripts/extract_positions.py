"""Leistungsverzeichnisse aus den Vergabeunterlagen → ``doc_positions.parquet``.

**Die Lücke.** `docpipe` zieht aus den Archiven nur TEXT. Ein GAEB-Leistungsverzeichnis ist
aber bereits strukturiert — Ordnungszahl, Menge, Einheit, Kurztext je Position. Diese Struktur
als Fließtext zu behandeln, wirft genau das weg, was den Wert ausmacht: „wie viel wovon".
`govisor.docparse.parse_gaeb` kann das seit Ticket 23, wurde aber nur beim EINZEL-UPLOAD
aufgerufen. Über den Korpus lief es nie.

**Was hier herauskommt** (je Position eine Zeile):
``notice_id, quelle, datei, rno, menge, einheit, text`` — plus eine Aggregatzeile je Vorgang
in ``doc_lv.parquet`` (Positionszahl, Summe der Mengen je Einheit als JSON).

**Grenzen, ehrlich.** GAEB DA XML (X8x) wird geparst; die alten Flat-Formate D8x nicht — die
sind ein Zeilenformat mit fester Spaltenbreite und brauchen einen eigenen Leser. XLSX liefert
nur Struktur (Blätter, Spaltenüberschriften, Zeilenzahl), bewusst KEINE Zellwerte: ein
Excel-LV hat kein festes Schema, und geratene Spaltenzuordnungen erzeugen falsche Mengen —
schlimmer als keine.

Aufruf:  python3 scripts/extract_positions.py [--country DE] [--limit N]
"""
from __future__ import annotations

import argparse
import json
import sys
import io
import re
import zipfile
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from govisor import docparse  # noqa: E402

_MAX_ENTPACKT = 60 * 1024 * 1024      # je Datei — Zip-Bomben-Schutz (wie docpipe)


# ── Preisblatt (Nicht-Bau) ────────────────────────────────────────────────────────────────────
# Was GAEB fuer den Bau ist, ist das Preisblatt fuer IT, Beratung, Reinigung, Medizin. Es hat
# KEIN einheitliches Schema — Zeile 0 traegt meist den Dokumenttitel, der echte Kopf steht
# irgendwo bis Zeile ~15. Gemessen an 22 Blaettern: 11 haben eine erkennbare Kopfzeile, 11
# nicht (Fliesstext-Formulare, „Summe der Pflegepauschalen" o. Ae.). Die halbe Ausbeute ist
# der ehrliche Erwartungswert, kein Defekt.
#
# WICHTIG: die PREIS-Spalten bleiben ungelesen — sie sind leer. Das Preisblatt ist ein
# Formular, das der Bieter ausfuellt. Uns interessiert die linke Haelfte: Position,
# Bezeichnung, Menge, Einheit. Das ist der Leistungsumfang.
_PB_ROLLEN = {
    "rno":     r"^(pos\.?|position|ordnungszahl|lfd)",
    "text":    r"bezeichnung|leistung|beschreibung|artikel|gegenstand",
    "menge":   r"^(ca\.?\s*)?menge|anzahl|stück|stueck",
    "einheit": r"^einheit(?!spreis)|^me$|mengeneinheit",
}
_PB_KOPF = re.compile(r"\b(pos|position|ordnungszahl|bezeichnung|leistung|menge|einheit|"
                      r"einheitspreis|gesamtpreis)\b", re.I)


def _pb_kopfzeile(ws):
    """Kopfzeile eines Preisblatts finden: >=2 LV-typische Begriffe in einer Zeile."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 15:
            return None
        zellen = [str(c).strip() if c is not None else "" for c in row]
        gefuellt = [c for c in zellen if c]
        if len(gefuellt) >= 3 and sum(1 for c in gefuellt if _PB_KOPF.search(c)) >= 2:
            rollen: dict[str, int] = {}
            for rolle, muster in _PB_ROLLEN.items():
                for j, c in enumerate(zellen):
                    if c and re.search(muster, c.replace("\n", " "), re.I):
                        rollen.setdefault(rolle, j)
            return i, rollen
    return None


def lies_preisblatt(daten: bytes, blatt: str) -> list[dict]:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(daten), data_only=True)
    except Exception:
        return []
    if blatt not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[blatt]
    kopf = _pb_kopfzeile(ws)
    if not kopf or "text" not in kopf[1]:
        wb.close()
        return []                                  # ohne Bezeichnungsspalte kein Leistungsumfang
    start, rollen = kopf
    aus, leer = [], 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= start:
            continue
        z = [str(c).strip() if c is not None else "" for c in row]

        def f(rolle):
            j = rollen.get(rolle)
            return z[j] if j is not None and j < len(z) else ""

        text = f("text")
        if not text:
            leer += 1
            if leer > 8:                           # Ende der Tabelle (Summenblock o. Ae.)
                break
            continue
        leer = 0
        aus.append({"rno": f("rno") or None, "qty": f("menge"),
                    "unit": f("einheit") or None, "text": text[:300]})
    wb.close()
    return aus


def _archive(vorgang: Path):
    """Alle ZIPs eines Vorgangs (in der Regel genau eines)."""
    return sorted(vorgang.glob("*.zip"))


# ⚠ HOCHZAEHLEN, WENN SICH AM PARSEN ETWAS AENDERT. Der Merker unten ueberspringt alles,
# was seit dem letzten Lauf unveraendert ist — er kennt aber nur die ARCHIVE, nicht den
# Code. Ein verbesserter GAEB- oder XLSX-Leser wuerde deshalb lautlos nie angewandt: die
# alten Ergebnisse blieben stehen und saehen frisch aus. Diese Zahl im Fingerabdruck macht
# aus einer Parser-Aenderung eine Neuberechnung.
PARSER_STAND = 1


def _fingerabdruck(vorgang: Path) -> str:
    """Woran man erkennt, dass sich an einem Vorgang nichts geaendert hat.

    Anzahl, Groesse und juengste Aenderung seiner Archive — plus `PARSER_STAND`. Kein Hash
    ueber den Inhalt: der kostet genau das Lesen, das hier gespart werden soll.
    """
    zs = _archive(vorgang)
    teile = [f"{z.name}:{st.st_size}:{int(st.st_mtime)}"
             for z in zs if (st := z.stat())]
    return f"v{PARSER_STAND}|" + "|".join(teile)


def _stand_lesen(root: Path) -> dict[str, str]:
    """Fingerabdruecke des letzten Laufs. Leer, wenn es keinen gab."""
    p = root / "doc_positions_stand.parquet"
    if not p.exists():
        return {}
    import duckdb
    try:
        return {str(a): str(b) for a, b in duckdb.connect().execute(
            f"select notice_id, fingerabdruck from read_parquet('{p.as_posix()}')").fetchall()}
    except Exception:
        return {}


def _alte_zeilen(root: Path) -> tuple[dict[str, list], dict[str, dict], bool]:
    """Positionen und LV-Zeilen des letzten Laufs, nach Vorgang sortiert.

    ⚠ OHNE DIESE ZEILEN DARF NICHTS UEBERSPRUNGEN WERDEN. Der Merker sagt nur „unveraendert";
    die Ergebnisse muessen trotzdem aus dem letzten Lauf uebernommen werden, sonst schrumpft
    die Ausgabe bei jedem Lauf um alles, was gerade nicht neu gerechnet wurde.
    """
    import duckdb
    con = duckdb.connect()
    pos: dict[str, list] = {}
    lv: dict[str, dict] = {}
    pp, lp = root / "doc_positions.parquet", root / "doc_lv.parquet"
    vorhanden = pp.exists() and lp.exists()
    if pp.exists():
        for r in con.execute(f"select * from read_parquet('{pp.as_posix()}')").fetchall():
            pos.setdefault(str(r[0]), []).append(
                {"notice_id": r[0], "quelle": r[1], "datei": r[2], "rno": r[3],
                 "menge": r[4], "einheit": r[5], "text": r[6]})
    if lp.exists():
        for r in con.execute(f"select * from read_parquet('{lp.as_posix()}')").fetchall():
            lv[str(r[0])] = {"notice_id": r[0], "n_positionen": r[1],
                             "mengen_je_einheit": r[2], "xlsx_blaetter": r[3]}
    return pos, lv, vorhanden


def sammle(country: str, limit: int | None,
           voll: bool = False) -> tuple[list[dict], list[dict], dict[str, str], tuple[int, int]]:
    """Leistungsverzeichnisse aller Vorgaenge — unveraenderte uebernommen statt neu gelesen.

    ⚠ WARUM UEBERHAUPT. Bis zum 2026-09-03 entpackte dieser Schritt JEDE Nacht alle
    Archive neu, auch die von vorgestern. Gemessen ueber 23 Nachtlaeufe: der Bestand wuchs
    um 187 %, die Dauer schwankte dabei um den Faktor 14 (357 s bis 4.999 s) — bei
    IDENTISCHEM Bestand von 2.316 Vorgaengen einmal 434 s und einmal 1.970 s. Die Zeit haengt
    also kaum am Umfang, sondern daran, ob gleichzeitig jemand dieselbe Platte benutzt; und
    `data` ist ein Symlink auf ein externes Volume, auf dem die beiden Dokument-Arbeiter
    rund um die Uhr schreiben. Wer weniger liest, streitet weniger.

    ⚠ „NICHTS GEFUNDEN" IST AUCH EIN ERGEBNIS. Nur 4.011 der 10.216 Vorgaenge haben
    ueberhaupt ein Leistungsverzeichnis. Wuerde der Merker nur die Treffer kennen, liefe der
    Schritt fuer die anderen 6.200 jede Nacht erneut — also fuer die Mehrheit. Deshalb
    haelt `doc_positions_stand.parquet` einen Fingerabdruck fuer JEDEN geprueften Vorgang.
    """
    root = ROOT / "data" / "docs" / country
    vorgaenge = sorted(p for p in root.iterdir() if p.is_dir())
    if limit:
        vorgaenge = vorgaenge[:limit]
    stand = {} if voll else _stand_lesen(root)
    alt_pos, alt_lv, alt_da = ({}, {}, False) if voll else _alte_zeilen(root)
    positionen: list[dict] = []
    lv: list[dict] = []
    neuer_stand: dict[str, str] = {}
    uebernommen = gelesen = 0
    for v in vorgaenge:
        nid = v.name
        fa = _fingerabdruck(v)
        neuer_stand[nid] = fa
        # ⚠ GEPRUEFT UND NICHTS GEFUNDEN IST AUCH EIN ERGEBNIS — und der haeufigste Fall:
        # nur 4.011 der 10.216 Vorgaenge haben ueberhaupt ein Leistungsverzeichnis. Die
        # Bedingung darf deshalb NICHT verlangen, dass der Vorgang in der alten Ausgabe
        # steht; sonst laesst sie genau die 6.200 durch, um die es beim Sparen geht. Der
        # erste Entwurf tat das und haette fast nichts gespart.
        #
        # Was sie stattdessen verlangt: die Ausgabedateien muessen ueberhaupt da sein.
        # Fehlen sie (geloescht, verschoben), gibt es nichts zu uebernehmen, und der
        # Merker allein wuerde die Zeilen fuer immer verschwinden lassen.
        if alt_da and stand.get(nid) == fa:
            positionen.extend(alt_pos.get(nid, []))
            if nid in alt_lv:
                lv.append(alt_lv[nid])
            uebernommen += 1
            continue
        gelesen += 1
        n_pos = 0
        mengen: Counter = Counter()
        blaetter: list[dict] = []
        gesehen: set = set()
        for z in _archive(v):
            try:
                zf = zipfile.ZipFile(z)
            except Exception:
                continue
            with zf:
                for info in zf.infolist():
                    if info.is_dir() or info.file_size > _MAX_ENTPACKT:
                        continue
                    endung = Path(info.filename).suffix.lower()
                    if endung not in docparse.GAEB_EXTS and endung not in (".xlsx", ".xlsm"):
                        continue
                    try:
                        daten = zf.read(info)
                    except Exception:
                        continue
                    if endung in docparse.GAEB_EXTS:
                        res = docparse.parse_gaeb(daten)
                        if not res:
                            continue          # D8x-Flatformat o. Ä. — ehrlich uebersprungen
                        for p in res["positions"]:
                            positionen.append({
                                # `res["parser"]` unterscheidet "gaeb" (DA XML) von "gaeb-flat"
                                # (DA 90). Ohne diese Trennung laesst sich der Beitrag des
                                # Flat-Lesers nicht mehr messen — beides hiesse nur "gaeb".
                                "notice_id": nid, "quelle": res["parser"], "datei": info.filename,
                                "rno": p.get("rno") or None,
                                "menge": _zahl(p.get("qty")),
                                "einheit": (p.get("unit") or None),
                                "text": (p.get("text") or None),
                            })
                            n_pos += 1
                            if p.get("unit") and _zahl(p.get("qty")) is not None:
                                mengen[p["unit"]] += _zahl(p["qty"])
                    else:
                        res = docparse.parse_xlsx(daten)
                        if not res:
                            continue
                        for sh in res.get("sheets", []):
                            blaetter.append({"datei": info.filename, **sh})
                            n_pos += sh.get("n_positions", 0)
                            # Preisblatt = Leistungsumfang der Nicht-Bau-Branchen.
                            if "preis" not in (sh.get("name") or "").lower():
                                continue
                            for p in lies_preisblatt(daten, sh["name"]):
                                schluessel = (nid, p.get("rno"), p.get("text"))
                                if schluessel in gesehen:
                                    continue          # gleiches Blatt in zwei Archivdateien
                                gesehen.add(schluessel)
                                positionen.append({
                                    "notice_id": nid, "quelle": "preisblatt",
                                    "datei": info.filename, "rno": p.get("rno"),
                                    "menge": _zahl(p.get("qty")), "einheit": p.get("unit"),
                                    "text": p.get("text"),
                                })
        if n_pos or blaetter:
            lv.append({
                "notice_id": nid,
                "n_positionen": n_pos,
                "mengen_je_einheit": json.dumps({k: round(v, 2) for k, v in mengen.most_common(12)},
                                                ensure_ascii=False) if mengen else None,
                "xlsx_blaetter": json.dumps(blaetter[:20], ensure_ascii=False) if blaetter else None,
            })
    return positionen, lv, neuer_stand, (uebernommen, gelesen)


def _zahl(s) -> float | None:
    if s in (None, ""):
        return None
    try:
        return float(str(s).replace(".", "").replace(",", ".")) if "," in str(s) else float(s)
    except ValueError:
        return None


def main(country: str, limit: int | None, voll: bool = False) -> int:
    import pyarrow as pa
    import pyarrow.parquet as pq

    positionen, lv, stand, (uebernommen, gelesen) = sammle(country, limit, voll)
    root = ROOT / "data" / "docs" / country
    if positionen:
        pq.write_table(pa.Table.from_pylist(positionen, schema=pa.schema([
            ("notice_id", pa.string()), ("quelle", pa.string()), ("datei", pa.string()),
            ("rno", pa.string()), ("menge", pa.float64()), ("einheit", pa.string()),
            ("text", pa.string())])), root / "doc_positions.parquet", compression="zstd")
    if lv:
        pq.write_table(pa.Table.from_pylist(lv, schema=pa.schema([
            ("notice_id", pa.string()), ("n_positionen", pa.int64()),
            ("mengen_je_einheit", pa.string()), ("xlsx_blaetter", pa.string())])),
            root / "doc_lv.parquet", compression="zstd")
    # ⚠ DER MERKER ZULETZT. Stirbt der Lauf zwischen den Ergebnissen und ihm, ist der
    # Merker aelter als die Daten — dann wird beim naechsten Mal zu viel gelesen, was
    # Zeit kostet, aber nichts kaputt macht. Andersherum waere es ein Datenverlust:
    # ein Merker ohne die zugehoerigen Zeilen laesst sie fuer immer uebersprungen.
    #
    # ⚠ UND NUR, WENN DIE ERGEBNISSE AUCH GESCHRIEBEN WURDEN. Bei `--limit` sieht der Lauf
    # nur einen Ausschnitt; einen Merker daraus zu schreiben hiesse, den Rest als geprueft
    # zu markieren, ohne ihn angesehen zu haben.
    if stand and limit is None:
        pq.write_table(pa.Table.from_pylist(
            [{"notice_id": k, "fingerabdruck": v} for k, v in sorted(stand.items())],
            schema=pa.schema([("notice_id", pa.string()), ("fingerabdruck", pa.string())])),
            root / "doc_positions_stand.parquet", compression="zstd")
    mit_gaeb = len({p["notice_id"] for p in positionen})
    print(f"Leistungsverzeichnisse {country}: {len(positionen):,} Positionen aus {mit_gaeb} Vorgängen "
          f"(GAEB), {len(lv)} Vorgänge mit LV insgesamt")
    print(f"  {gelesen:,} Vorgänge gelesen, {uebernommen:,} unverändert übernommen"
          + ("  (--voll: alles gelesen)" if voll else ""))
    return 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--country", default="DE")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--voll", action="store_true",
                    help="alles neu lesen statt Unveraendertes zu uebernehmen")
    a = ap.parse_args()
    sys.exit(main(a.country, a.limit, a.voll))
