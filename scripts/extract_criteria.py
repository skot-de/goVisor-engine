"""Kriterien-Matrizen (UfAB/EVB-IT) aus XLSX → ``doc_criteria.parquet``.

**Warum das der wichtigste noch fehlende Baustein war.** Die Parser-Selbstdiagnose wies für
``award_weights`` eine Lücke von 102 aus, und die häufigste unerkannte Textform lautete
*„Mehrere Zuschlagskriterien gemäß Formblatt"* — die Gewichte stehen also nicht im Fließtext,
sondern in einem separaten Formblatt. Das Formblatt liegt den Unterlagen bei: als Excel mit
stabilem Spaltenschema (Kriterienhauptgruppe / Kriteriengruppe / Kriterium / Gewichtung).

Und es trägt mehr als Gewichte: die Spalte hinter dem Kriteriums-Code führt **A oder B** —
``A = Ausschlusskriterium`` (nicht erfüllt ⇒ Angebot fliegt raus), ``B = Bewertungskriterium``
(bringt Punkte). Das ist die K.o.-Liste, also die erste Frage jedes Bieters.

**Warum hier Zellwerte gelesen werden.** ``docparse.parse_xlsx`` liest bewusst keine Werte
(§6.2). Diese Regel stammt aus dem Upload-Pfad, wo Nutzer eigene Dokumente hochladen. Hier
geht es um veröffentlichte Vergabeunterlagen von öffentlichen Portalen — von Sven ausdrücklich
freigegeben (2026-08-13). Der Upload-Pfad bleibt unverändert.

**Robustheit.** Spalten werden über die ÜBERSCHRIFT gefunden, nie über einen festen Index:
die Matrizen nutzen verbundene Zellen, dieselbe Spalte liegt je Formular woanders. Was nicht
erkannt wird, wird übersprungen — lieber ein Vorgang weniger als eine falsche Zuordnung.

Aufruf:  python3 scripts/extract_criteria.py [--country DE]
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
import warnings
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

warnings.filterwarnings("ignore", module="openpyxl")

# Überschriften, an denen die Matrix erkannt wird. Der Kern ist „Kriterienhauptgruppe" —
# ohne den ist es eine andere Tabelle und wird nicht angefasst.
_MARKER = "kriterienhauptgruppe"
_SPALTEN = {
    "khg":        r"kriterienhauptgruppe",
    "kg":         r"kriteriengruppe",
    "k":          r"^kriterium",
    "gewichtung": r"^gewichtung",
    "prozent":    r"^prozent",
    # Nur die zweite Bauform fuehrt sie; in der ersten steht die Art im Code selbst.
    "art":        r"^art$",
}
# Kriteriums-Codes kommen in mehreren Schreibweisen vor (gemessen 2026-08-24 ueber 18
# Matrizen). Zwei davon sind eindeutig und werden gelesen:
#
#   „A.1.1.4" / „B.2.3"      der Buchstabe IST die Art          (die urspruengliche Form)
#   „Kriterium A.1.2"        dasselbe, mit Wort davor
#   „K 1.1.1"                Nummer ohne Art; die steht in einer eigenen Spalte als „[ A ]"
#
# ⚠ Sechs der achtzehn tragen GAR KEINE erkennbare Code-Form. Sie bleiben ungelesen. Der
# Grundsatz dieses Skripts gilt weiter: lieber ein Vorgang weniger als eine falsche
# Zuordnung — eine erratene Ausschlussliste ist schlimmer als keine.
_CODE = re.compile(r"^(?:kriterium\s+)?([AB])\.((?:\d+\.)*\d+)$", re.I)
# Die zweite Form. Bewusst eng: ohne das fuehrende „K" wuerde jede Zahl in einer
# Gewichtungsspalte als Kriteriums-Code durchgehen.
_CODE_K = re.compile(r"^K[\s.]?((?:\d+\.)*\d+)$", re.I)
_ART_ZELLE = re.compile(r"^\[?\s*([AB])\s*\]?$", re.I)


def _ist_code(zelle: str) -> bool:
    """Sieht die Zelle aus wie ein Kriteriums-Code? Dann ist sie kein Gruppenname."""
    return bool(zelle) and bool(_CODE.match(zelle) or _CODE_K.match(zelle))
# Gruppenzeilen benennen sich in beiden Bauformen selbst.
_KHG_ZEILE = re.compile(r"^KHG\b", re.I)
_KG_ZEILE = re.compile(r"^KG\b", re.I)
_MAX_ZEILEN = 4000


def _kopfzeile(ws) -> tuple[int, dict[str, int]] | None:
    """Kopfzeile suchen und Spaltenüberschrift → Index abbilden (nicht über feste Positionen)."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 25:
            return None
        zellen = [str(c).strip().lower() if c is not None else "" for c in row]
        if not any(_MARKER in c for c in zellen):
            continue
        idx: dict[str, int] = {}
        for name, muster in _SPALTEN.items():
            for j, c in enumerate(zellen):
                if c and re.search(muster, c):
                    idx.setdefault(name, j)
        # „Gewichtung" ist eine VERBUNDENE Zelle ueber drei Unterspalten. Die Werte stehen
        # nicht darunter, sondern in den Unterspalten, die erst die naechste Zeile benennt:
        #   Zeile 0:  … | Gewichtung |    |    | Prozent | …
        #   Zeile 1:  … |    KHG     | KG | K  |         | …
        # Ohne diesen Schritt las ich die verbundene Kopfspalte und bekam ueberall None —
        # gemessen 0 von 6.832 Kriterien mit Gewichtung.
        # ⚠ Auch „Art" steht in manchen Matrizen erst in der Zeile UNTER der Kopfzeile —
        # dieselbe Bauart wie bei „Gewichtung". Ohne diesen Blick fehlt die Spalte, und
        # damit bleiben die Kriterien ungelesen: ohne belegte Art wird nichts geraten.
        unter_z = next(ws.iter_rows(min_row=i + 2, max_row=i + 2, values_only=True), ())
        unter_z = [str(c).strip().lower() if c is not None else "" for c in unter_z]
        if "art" not in idx:
            for j, c in enumerate(unter_z):
                if c == "art":
                    idx["art"] = j
                    break
        if "gewichtung" in idx:
            unter = unter_z
            j0 = idx["gewichtung"]
            for j in range(j0, min(j0 + 4, len(unter))):
                if unter[j] == "khg":
                    idx["gew_khg"] = j
                elif unter[j] == "kg":
                    idx["gew_kg"] = j
                elif unter[j] == "k":
                    idx["gew_k"] = j
        return i, idx
    return None


def _zahl(v) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def lies_datei(archiv: Path, datei: str) -> bytes | None:
    """Datei aus dem Archiv holen — auch aus einem Archiv IM Archiv.

    ⚠ Der Volltext-Index notiert verschachtelte Pfade mit `::`:
    ``Vergabeunterlagen/Version 3/Z42.zip::Anlage 510-3 Kriterienkatalog.xlsx``. Wer den
    Pfad unverändert an `ZipFile.read` gibt, bekommt `KeyError` — und der wurde bisher
    stillschweigend übersprungen. Zwei der zehn fehlenden Matrizen lagen genau so.
    """
    teile = datei.split("::")
    try:
        with zipfile.ZipFile(archiv) as zf:
            daten = zf.read(teile[0])
        for inneres in teile[1:]:
            with zipfile.ZipFile(io.BytesIO(daten)) as zf:
                daten = zf.read(inneres)
        return daten
    except Exception:                                    # noqa: BLE001
        return None


def lies_matrix(daten: bytes, blatt: str | None = None) -> list[dict]:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(daten), read_only=False, data_only=True)
    except Exception:
        return []
    # ⚠ NICHT nur EIN Blatt. Bis zum 2026-08-24 bekam diese Funktion den Blattnamen aus dem
    # Vorlauf vorgegeben — und der Marker sitzt fast nie dort, wo der Vorlauf ihn vermutete:
    # gemessen auf „Erklärung", „Übersicht", „Erläuterungen", „Bewertung VgU A-Kriterien".
    # Eine Datei kann ihn sogar auf ZWEI Blättern tragen („Qualität - Variante 1" und
    # „Variante 2"); beide sind echte Angaben und beide gehören gelesen.
    blaetter = [blatt] if blatt in wb.sheetnames else list(wb.sheetnames)
    aus: list[dict] = []
    for name in blaetter:
        aus.extend(_lies_blatt(wb[name], name))
    wb.close()
    return aus


def _lies_blatt(ws, blattname: str) -> list[dict]:
    kopf = _kopfzeile(ws)
    if not kopf:
        return []
    start, idx = kopf
    aus: list[dict] = []
    khg = kg = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= start + 1 or i > _MAX_ZEILEN:
            continue
        z = [str(c).strip() if c is not None else "" for c in row]

        def feld(name: str) -> str:
            j = idx.get(name)
            return z[j] if j is not None and j < len(z) else ""

        # Ebene 1/2: Hauptgruppe bzw. Gruppe merken — sie stehen einmal und gelten fuer die
        # darunter liegenden Kriterien weiter (typische Excel-Gliederung ohne Wiederholung).
        # ⚠ EIN GRUPPENNAME IST NIE EIN CODE. In der zweiten Bauform zeigen `khg` und `kg`
        # auf dieselbe Spalte, in der auch die Codes stehen — ohne Pruefung wuerde jede
        # Kriteriumszeile als neue Hauptgruppe gelesen („khg = K 1.1.1"). Geprueft wird am
        # WERT, nicht an der Spaltenanordnung: ein Riegel, der nur bei gleicher Spalte
        # greift, faellt schon bei der naechsten Variante wieder auf.
        roh_khg, roh_kg = feld("khg"), feld("kg")
        if _ist_code(roh_khg):
            roh_khg = ""
        if _ist_code(roh_kg):
            roh_kg = ""
        # Und beide benennen sich selbst: „KHG 1: …" ist keine Gruppe, „KG 1.1 …" keine
        # Hauptgruppe — auch dann nicht, wenn sie in derselben Spalte stehen.
        if _KG_ZEILE.match(roh_khg):
            roh_khg = ""
        if _KHG_ZEILE.match(roh_kg):
            roh_kg = ""
        if not roh_khg:
            roh_khg = next((c for c in z if _KHG_ZEILE.match(c)), "")
        if roh_khg:
            khg = roh_khg
            kg = None
        if roh_kg:
            # In der Gruppenzeile steht links die Nummer, rechts daneben der Name.
            j = idx["kg"]
            name = next((x for x in z[j + 1:j + 4] if x), "")
            kg = f"{roh_kg} {name}".strip()
        else:
            selbst = next((c for c in z if _KG_ZEILE.match(c)), "")
            if selbst:
                kg = selbst

        # Ebene 3: das eigentliche Kriterium — erkennbar am Code.
        code = txt = art = ""
        for j, c in enumerate(z):
            m = _CODE.match(c)
            if m:
                code, art = c, m.group(1).upper()
                txt = next((x for x in z[j + 1:j + 3] if len(x) > 3), "")
                break
        if not code and "art" in idx:
            # Zweite Bauform: „K 1.1.1" plus eine Spalte „Art" mit „[ A ]" bzw. „[ B ]".
            # NUR wenn diese Spalte existiert — ohne sie waere die Art geraten, und ein
            # falsch als Bewertungskriterium gefuehrter Ausschluss ist ein Schaden.
            for j, c in enumerate(z):
                if _CODE_K.match(c):
                    m_art = _ART_ZELLE.match(feld("art"))
                    if not m_art:
                        break
                    code, art = c, m_art.group(1).upper()
                    txt = next((x for x in z[j + 1:j + 4] if len(x) > 3
                                and not _ART_ZELLE.match(x)), "")
                    break
        if not code:
            continue
        aus.append({
            "blatt": blattname,
            "khg": khg or None, "kg": kg or None, "code": code,
            "art": "ausschluss" if art == "A" else "bewertung",
            "text": txt[:400] or None,
            # Gewicht auf Kriteriumsebene; die Gruppen-Gewichte stehen in den
            # Gruppenzeilen und gelten fuer alle Kriterien darunter.
            "gewichtung": _zahl(feld("gew_k")) or _zahl(feld("gewichtung")),
            "gew_gruppe": _zahl(feld("gew_kg")),
            "gew_hauptgruppe": _zahl(feld("gew_khg")),
            "prozent": _zahl(feld("prozent")),
        })
    return aus


def main(country: str) -> int:
    import duckdb
    import pyarrow as pa
    import pyarrow.parquet as pq

    root = ROOT / "data" / "docs" / country
    lv = root / "doc_lv.parquet"
    if not lv.exists():
        print("kein doc_lv.parquet — erst `extract_positions.py` laufen lassen.")
        return 1
    con = duckdb.connect()
    # ⚠ KANDIDATEN AUS DEM VOLLTEXT-INDEX, NICHT AUS `doc_lv`.
    #
    # Bis zum 2026-08-24 kamen sie aus `doc_lv.parquet` — dem Ergebnis der
    # Leistungsverzeichnis-Extraktion. Damit hing die Kriterien-Erkennung an einem
    # Vorlauf, der etwas ganz anderes sucht, und verlor auf zwei Wegen:
    #
    #   · 2 von 10 fehlenden Vorgaengen standen gar nicht in `doc_lv` — wer kein
    #     Leistungsverzeichnis hat, wurde nie auf Kriterien geprueft;
    #   · 7 von 8 uebrigen standen zwar drin, trugen den Marker aber nicht in
    #     `xlsx_blaetter`, weil dessen Spaltenerkennung ihn nicht erwischt hatte.
    #
    # `doc_text` weiss dagegen direkt, WELCHE Datei den Marker enthaelt. Gegengeprueft:
    # der neue Weg findet alle 8 bisher extrahierten Vorgaenge und 10 weitere.
    text_p = root / "doc_text.parquet"
    if not text_p.exists():
        print("kein doc_text.parquet — erst `index-docs` laufen lassen.")
        return 1
    kandidaten = con.execute(
        f"""SELECT DISTINCT notice_id, archive, file
            FROM read_parquet('{text_p.as_posix()}')
            WHERE lower(filetype) = '.xlsx' AND lower(text) LIKE '%{_MARKER}%'
            ORDER BY 1""").fetchall()

    zeilen: list[dict] = []
    ohne = 0
    for nid, archiv, datei in kandidaten:
        pfad = (root / nid) / archiv if archiv else None
        if pfad is None or not pfad.exists():
            pfad = next((root / nid).glob("*.zip"), None)
        if pfad is None:
            ohne += 1
            continue
        daten = lies_datei(pfad, datei)
        if daten is None:
            ohne += 1
            continue
        vorher = len(zeilen)
        for k in lies_matrix(daten):
            zeilen.append({"notice_id": nid, "datei": datei, **k})
        if len(zeilen) == vorher:
            ohne += 1

    if zeilen:
        pq.write_table(pa.Table.from_pylist(zeilen, schema=pa.schema([
            ("notice_id", pa.string()), ("datei", pa.string()), ("blatt", pa.string()),
            ("khg", pa.string()),
            ("kg", pa.string()), ("code", pa.string()), ("art", pa.string()),
            ("text", pa.string()), ("gewichtung", pa.float64()),
            ("gew_gruppe", pa.float64()), ("gew_hauptgruppe", pa.float64()),
            ("prozent", pa.float64())])),
            root / "doc_criteria.parquet", compression="zstd")
    n_vorgang = len({z["notice_id"] for z in zeilen})
    n_aus = sum(1 for z in zeilen if z["art"] == "ausschluss")
    print(f"Kriterien-Matrizen {country}: {len(zeilen):,} Kriterien aus {n_vorgang} Vorgängen "
          f"({n_aus:,} Ausschluss-, {len(zeilen)-n_aus:,} Bewertungskriterien)")
    if ohne:
        print(f"  {ohne} Vorgänge mit Matrix-Überschrift, aber ohne lesbare Zeilen — "
              f"abweichendes Layout, bewusst übersprungen statt geraten.")
    return 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--country", default="DE")
    sys.exit(main(ap.parse_args().country))
