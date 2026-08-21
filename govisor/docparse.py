"""Parser-Schiene für strukturierte Vergabeunterlagen (Ticket #23, §6.2).

Wo ein Format strukturiert vorliegt, ist der LLM überflüssig (und teuer): GAEB-LV (84 % im
Bau!), ausfüllbare PDF-Formulare und Excel-Preisblätter werden **regelbasiert** geparst.
Ausgabe ist strukturierte Fakten (Positionen/Felder/Tabellenstruktur) — bei Preisblättern
bewusst **ohne Werte** (§6.2: „keine Werte eintragen"). Wo ein Parser greift, entfällt die
LLM-Extraktion für dieses Dokument.

Zusätzlich der **Inhalts-Klassifikator** (§6.1, Schritt 2) für die 31 % Dateien, die der
Dateiname nicht trifft.
"""
from __future__ import annotations

import io
import re

# Endungen, für die ein Parser existiert (→ keine LLM-Extraktion).
GAEB_EXTS = frozenset({".x83", ".x81", ".x86", ".d83", ".d81", ".p83", ".gaeb"})
XLSX_EXTS = frozenset({".xlsx", ".xlsm"})


def _localname(tag) -> str:
    """XML-Tag ohne Namensraum. Nimmt AUCH Nicht-Strings entgegen.

    lxml liefert fuer Kommentare und Processing-Instructions als `.tag` eine FUNKTION,
    keinen String — `root.iter()` gibt diese Knoten mit aus. Ohne diesen Guard warf der
    GAEB-Parser dort `TypeError: argument of type 'cython_function_or_method' is not a
    container`. Aufgefallen erst beim Lauf ueber echte Leistungsverzeichnisse; beim
    Einzel-Upload hatte es nie jemanden getroffen.
    """
    if not isinstance(tag, str):
        return ""
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def parse_gaeb(data: bytes) -> dict | None:
    """GAEB → Positionen (Ordnungszahl, Menge, Einheit, Kurztext).

    Zuerst DA XML (X8x), namespace-agnostisch über local-name (deckt DA83 3.2/3.3 ab). Ist es
    kein XML, greift ``parse_gaeb_flat`` für die alten Zeilenformate D81/D83/P83.
    """
    try:
        from lxml import etree
        root = etree.fromstring(data)
    except Exception:
        return parse_gaeb_flat(data)
    positions = []
    for el in root.iter():
        if _localname(el.tag) != "Item":
            continue
        rno = el.get("RNoPart") or el.get("RNo") or ""
        qty = qu = short = ""
        for ch in el.iter():
            ln = _localname(ch.tag)
            t = (ch.text or "").strip()
            if ln == "Qty" and not qty:
                qty = t
            elif ln == "QU" and not qu:
                qu = t
            elif ln in ("OutlineText", "TextOutlTxt", "span") and t and not short:
                short = t
        if not short:                       # Fallback: erster Text unter Description
            desc = next((c for c in el.iter() if _localname(c.tag) == "Description"), None)
            if desc is not None:
                short = " ".join(x.strip() for x in desc.itertext() if x.strip())[:200]
        if rno or qty or short:
            positions.append({"rno": rno, "qty": qty, "unit": qu, "text": short[:300]})
    if not positions:
        return None
    return {"parser": "gaeb", "positions": positions, "n_positions": len(positions)}


# ── GAEB DA 90: die alten Zeilenformate D81/D83/P83 ────────────────────────────────────────
#
# Kein XML, sondern feste Spaltenbreiten mit zweistelligem Satztyp am Zeilenanfang. Gemessen an
# unserem Bestand: 510 solcher Dateien, davon 144 Vorgänge, die AUSSCHLIESSLICH so vorliegen —
# ohne diesen Leser haben sie gar kein maschinenlesbares Leistungsverzeichnis.
#
# Der Positionssatz, an echten Dateien abgelesen (beide Ausprägungen kommen vor):
#
#     2101     1 NNN         00000001000Psch
#     2111 1     NNN         00000600000qm
#     ^^|-- OZ --|^^^|       |-- Menge -|^^^^
#     0 2         11 14      23         34
#
#   [2:11]   Ordnungszahl, 9 Zeichen, hierarchisch gruppiert
#   [11:14]  Kennzeichen (Bedarfs-/Alternativposition o. Ä.) — hier nicht ausgewertet
#   [23:34]  Menge, 11 Ziffern mit DREI IMPLIZITEN Nachkommastellen (00000600000 = 600,0)
#   [34:38]  Einheit
#
# Der Kurztext steht im Satz 25 dahinter — aber NICHT immer unmittelbar: dazwischen können
# Sätze 27 o. Ä. liegen. Deshalb wird bis zur nächsten Position vorwärts gesammelt statt die
# Folgezeile zu nehmen.
#
# Die OZ-Maske (welche Stellen Titel, welche Position sind) steht im Kopfsatz und wird hier
# NICHT gelesen. Die Gruppierung nach Leerraum ist eine Rekonstruktion — für die Anzeige
# ausreichend, als Sortierschlüssel nicht belastbar.
_FLAT_MENGE = slice(23, 34)
_FLAT_EINHEIT = slice(34, 38)
_FLAT_OZ = slice(2, 11)


def _flat_text(data: bytes) -> str | None:
    """DOS-Kodierung entschlüsseln. GAEB DA 90 stammt aus der DOS-Zeit — cp850 zuerst."""
    for enc in ("cp850", "cp437", "cp1252"):
        try:
            t = data.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
        # Ein DA-90-Kopf beginnt mit Satz 00; ohne den ist es eine andere Datei.
        if t.lstrip().startswith("00"):
            return t
    return None


def parse_gaeb_flat(data: bytes) -> dict | None:
    """GAEB DA 90 (D81/D83/P83) → dieselbe Struktur wie ``parse_gaeb``."""
    t = _flat_text(data)
    if t is None:
        return None
    positions: list[dict] = []
    offen: dict | None = None

    def schliesse():
        if offen and (offen["rno"] or offen["qty"] or offen["text"]):
            offen["text"] = offen["text"][:300]
            positions.append(offen)

    for zeile in t.splitlines():
        if len(zeile) < 3:
            continue
        satz = zeile[:2]
        if satz == "21":
            schliesse()
            roh = zeile[_FLAT_MENGE].strip()
            # Nur Ziffern sind eine Menge. Steht dort etwas anderes, ist die Zeile kürzer als
            # erwartet oder anders belegt — dann lieber keine Menge als eine erfundene.
            menge = f"{int(roh) / 1000:g}" if roh.isdigit() else ""
            offen = {
                "rno": " ".join(zeile[_FLAT_OZ].split()).replace(" ", "."),
                "qty": menge,
                "unit": zeile[_FLAT_EINHEIT].strip(),
                "text": "",
            }
        elif satz == "25" and offen is not None:
            # Jede DA-90-Zeile ist 80 Zeichen breit und traegt in [74:80] eine laufende
            # Zeilennummer. Wer nur `zeile[2:]` nimmt, haengt sie an den Text: aus
            # „Abwasserrohr DN100" wird „Abwasserrohr DN100    000138". Faellt in der
            # Anzeige kaum auf und steht doch in jeder zweiten Position.
            stueck = (zeile[2:74] if len(zeile) >= 80 and zeile[74:80].isdigit()
                      else zeile[2:]).strip()
            if stueck:
                offen["text"] = (offen["text"] + " " + stueck).strip()
        elif satz in ("11", "12"):        # Titel-/Gliederungszeile beendet die offene Position
            schliesse()
            offen = None
    schliesse()
    if not positions:
        return None
    return {"parser": "gaeb-flat", "positions": positions, "n_positions": len(positions)}


def parse_pdf_fields(data: bytes) -> dict | None:
    """Ausfüllbares PDF → Formularfelder (Name, Typ, Pflichtkennzeichen). None ohne Felder."""
    try:
        import pypdf
        r = pypdf.PdfReader(io.BytesIO(data))
        fields = r.get_fields()
    except Exception:
        return None
    if not fields:
        return None
    out = []
    for name, f in fields.items():
        try:
            ff = f.get("/Ff")
            required = bool(int(ff) & 2) if ff is not None else False   # Bit 2 = Required
        except Exception:
            required = False
        out.append({"name": str(name), "type": str(f.get("/FT", "")).lstrip("/") or "text",
                    "required": required})
    return {"parser": "pdf_fields", "fields": out, "n_fields": len(out)}


def parse_xlsx(data: bytes) -> dict | None:
    """XLSX → Tabellenstruktur (Blätter, Spaltenüberschriften, Positionszahl). KEINE Werte (§6.2)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None
    sheets = []
    for ws in wb.worksheets:
        # ws.max_row ist im read_only-Modus unzuverlässig (deklarierte Dimension, oft 1.048.576).
        # Echte Positionszahl = nicht-leere Datenzeilen, gezählt beim Iterieren (gedeckelt).
        header, n_rows = [], 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [c for c in row if c is not None and str(c).strip()]
            if i == 0:
                header = [str(c).strip() for c in cells][:30]
                continue
            if cells:
                n_rows += 1
            if i > 20000:                                # Deckel gegen pathologische Blätter
                break
        sheets.append({"name": ws.title, "columns": header, "n_positions": n_rows})
    wb.close()
    if not sheets:
        return None
    return {"parser": "xlsx", "sheets": sheets}


def parse(name: str, ext: str, data: bytes) -> dict | None:
    """Dispatcher: strukturierte Ausgabe je Format, sonst None (→ LLM-/Text-Schiene)."""
    e = (ext or "").lower()
    if e in GAEB_EXTS:
        return parse_gaeb(data)
    if e in XLSX_EXTS:
        return parse_xlsx(data)
    if e == ".pdf":
        return parse_pdf_fields(data)
    return None


def classify_content(text: str, sample: int = 4000) -> str:
    """Doktyp aus einer Inhaltsprobe — liegt jetzt in :mod:`govisor.doctypes`.

    ⚠ Hier stand bis 2026-08-21 ein eigener, ungemessener Regelsatz nach dem Muster
    Erster-Treffer-gewinnt. Gegen eine Rueckhaltestichprobe gehalten lag er in ueber der
    Haelfte der Faelle daneben (46,1 % Genauigkeit). Die Nachfolge wertet nach Punkten und
    kommt auf 86 %. Diese Funktion bleibt als Aufrufname bestehen.
    """
    from . import doctypes
    return doctypes.classify_content(text, sample=sample)
